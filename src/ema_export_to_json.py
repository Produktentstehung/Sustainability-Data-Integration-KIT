#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Copyright (c) 2025 Heinz Nixdorf Institute
# Copyright (c) 2025 Paderborn University
# Copyright (c) 2025 Contributors to the Eclipse Foundation
#
# See the NOTICE file(s) distributed with this work for additional
# information regarding copyright ownership.
#
# This program and the accompanying materials are made available under the
# terms of the Apache License, Version 2.0 which is available at
# https://www.apache.org/licenses/LICENSE-2.0.
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS, WITHOUT
# WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
# License for the specific language governing permissions and limitations
# under the License.
#
# SPDX-License-Identifier: Apache-2.0
"""Reads an export of the ema Plant Designer simulation and writes the
process-related data as JSON to standard output.

Two sheets of the export are evaluated:

* "<year> - Durchlaufzeiten"  - operations with cycle time, processing time
                                and piece count
* "<year> - Medienverbrauch"  - energy demand per workstation and period

The energy demand is given per workstation, not per operation. It is therefore
distributed across the operations of that workstation in proportion to their
processing time, and related to a single piece.

The sheet names are those of the German user interface of ema Plant Designer;
they are used as they appear in the export.

Usage:  python ema_export_to_json.py <export.xlsx> [scenario]
        The scenario is the column identifier of the export, default: the
        first column
"""
import json
import re
import sys
import warnings

warnings.filterwarnings("ignore", module="openpyxl")

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  python -m pip install openpyxl")

# kWh -> MJ; openLCA fuehrt den Strombezug in MJ
KWH_TO_MJ = 3.6


def zahl(wert, standard=0.0):
    """Robuste Zahlwandlung. Exporte enthalten teils Platzhalter wie '...'
    oder Zahlen mit Komma als Dezimaltrenner."""
    if wert is None:
        return standard
    if isinstance(wert, (int, float)):
        return float(wert)
    text = str(wert).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return standard


def find_sheet(wb, suffix):
    for name in wb.sheetnames:
        if name.strip().lower().endswith(suffix.lower()):
            return wb[name]
    return None


def header_index(rows, label, scenario=None):
    """Sucht die Spalte zu einer Ueberschrift. Die Exporte fuehren die
    Szenariokennung (z. B. E8) in einer eigenen Zeile unter der Ueberschrift."""
    head, sub = rows[0], rows[1] if len(rows) > 1 else ()
    treffer = [i for i, h in enumerate(head)
               if h and label.lower() in str(h).lower()]
    if not treffer:
        return None
    if scenario:
        for i in treffer:
            if i < len(sub) and sub[i] and str(sub[i]).strip() == scenario:
                return i
    return treffer[0]


def parse_operations(ws, scenario):
    rows = [r for r in ws.iter_rows(values_only=True)]
    start = next((i for i, r in enumerate(rows)
                  if any(c and "Arbeitsgangnummer" in str(c) for c in r)), None)
    if start is None:
        return []
    block = rows[start:]
    idx = {
        "nummer": header_index(block, "Arbeitsgangnummer"),
        "name": header_index(block, "Arbeitsgangname"),
        "platz": header_index(block, "Arbeitsplatzname", scenario),
        "te": header_index(block, "Te [Sek]", scenario),
        "bz": header_index(block, "Bearbeitungszeit (BZ)", scenario),
        "stueck": header_index(block, "IST-Stückzahl", scenario),
    }
    if idx["nummer"] is None or idx["name"] is None:
        return []

    # Der ema-Export kuerzt Wiederholungen mit "..." ab. Diese Zellen
    # take the last value read from the same column.
    WIEDERHOLUNG = {"...", "…", '"'}
    letzte = {}

    def wert(row, spalte, schluessel):
        if spalte is None or spalte >= len(row):
            return 0.0
        roh = row[spalte]
        if roh is not None and str(roh).strip() in WIEDERHOLUNG:
            return letzte.get(schluessel, 0.0)
        z = zahl(roh)
        letzte[schluessel] = z
        return z

    ops = []
    for r in block[2:]:
        nr = r[idx["nummer"]] if idx["nummer"] < len(r) else None
        if nr is None or not str(nr).strip():
            continue
        platz = r[idx["platz"]] if idx["platz"] is not None and idx["platz"] < len(r) else ""
        platz = re.sub(r"\s*\(.*\)$", "", str(platz or "")).strip()
        ops.append({
            "operationNumber": str(nr).strip(),
            "operationName": str(r[idx["name"]] or "").strip(),
            "workstation": platz,
            "cycleTimeSeconds": wert(r, idx["te"], "te"),
            "processingTimeHours": wert(r, idx["bz"], "bz"),
            "quantity": wert(r, idx["stueck"], "stueck"),
        })
    return ops


def parse_media(ws, scenario):
    rows = [r for r in ws.iter_rows(values_only=True)]
    start = next((i for i, r in enumerate(rows)
                  if any(c and "Name des Arbeitsplatzes" in str(c) for c in r)), None)
    if start is None:
        return {}
    head = rows[start]
    medium = rows[start + 1] if len(rows) > start + 1 else ()
    scen = rows[start + 2] if len(rows) > start + 2 else ()

    name_idx = next(i for i, h in enumerate(head) if h and "Name des Arbeitsplatzes" in str(h))
    spalte = None
    for i, h in enumerate(head):
        if not h or "Verbrauch" not in str(h):
            continue
        if scenario and i < len(scen) and scen[i] and str(scen[i]).strip() != scenario:
            continue
        spalte = i
        break
    if spalte is None:
        return {}

    einheit = None
    for c in medium:
        if c and str(c).strip():
            einheit = str(c).strip()
            break

    werte = {}
    for r in rows[start + 3:]:
        if name_idx >= len(r) or not r[name_idx]:
            continue
        name = re.sub(r"\s*\(.*\)$", "", str(r[name_idx])).strip()
        wert = r[spalte] if spalte < len(r) else None
        if wert is None:
            continue
        werte[name] = {"amount": zahl(wert), "medium": einheit or "Strom"}
    return werte


def main():
    # Siehe repair_aasx.py: "--help" darf nicht als Dateiname enden.
    if len(sys.argv) < 2 or sys.argv[1] in ('--help', '-h', '/?'):
        print((__doc__ or '').strip()
              or 'Usage: python ema_export_to_json.py <export.xlsx> [scenario]')
        return
    pfad = sys.argv[1]
    scenario = sys.argv[2] if len(sys.argv) > 2 else None

    wb = openpyxl.load_workbook(pfad, data_only=True)

    ws_zeit = find_sheet(wb, "Durchlaufzeiten")
    ws_medien = find_sheet(wb, "Medienverbrauch")
    if ws_zeit is None:
        sys.exit("sheet 'Durchlaufzeiten' not found in the export")

    if scenario is None:
        rows = list(ws_zeit.iter_rows(values_only=True))
        start = next((i for i, r in enumerate(rows)
                      if any(c and "Arbeitsgangnummer" in str(c) for c in r)), 0)
        sub = rows[start + 1] if len(rows) > start + 1 else ()
        scenario = next((str(c).strip() for c in sub if c and str(c).strip()), None)

    ops = parse_operations(ws_zeit, scenario)
    medien = parse_media(ws_medien, scenario) if ws_medien else {}

    # Energie je Arbeitsplatz anteilig nach Bearbeitungszeit auf die
    # distribute across the operations and relate to a single piece
    zeit_je_platz = {}
    for op in ops:
        zeit_je_platz[op["workstation"]] = zeit_je_platz.get(op["workstation"], 0.0) \
            + op["processingTimeHours"]

    for op in ops:
        platz = op["workstation"]
        info = medien.get(platz)
        stueck = op["quantity"] or 1.0
        gesamt = zeit_je_platz.get(platz) or 0.0
        if info and gesamt > 0:
            anteil = op["processingTimeHours"] / gesamt
            kwh_je_stueck = info["amount"] * anteil / stueck
        elif info:
            kwh_je_stueck = info["amount"] / stueck
        else:
            kwh_je_stueck = 0.0
        op["energyPerUnitKWh"] = kwh_je_stueck
        op["energyPerUnitMJ"] = kwh_je_stueck * KWH_TO_MJ
        op["energySource"] = info["medium"] if info else None

    print(json.dumps({
        "scenario": scenario,
        "workstations": [{"name": k, "consumption": v["amount"], "unit": "kWh/Periode",
                          "medium": v["medium"]} for k, v in medien.items()],
        "operations": ops,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
